Расчет фин. эффекта (новый)¶

_df_effect = X_test_freq.copy()
_df_effect.shape

def payments_fee(row):
    payments = 0
    if row['Сумма_взыскано_по_ФУ'] > 0:
        payments += 100000
#    if row['Суммы_взыскано_по_иску'] > 0:
#        payments += 15000

    return payments

_df_effect['Взносы'] = _df_effect.apply(payments_fee, axis=1)
_df_effect['Взносы'].value_counts(dropna=False)

_df_effect['Сумма_выплат_по_претензиям'] = _df_effect['Сумма_выплат_по_претензиям'].fillna(0)
_df_effect['Сумма_взыскано_по_ФУ'] = _df_effect['Сумма_взыскано_по_ФУ'].fillna(0)
_df_effect['Суммы_взыскано_по_иску'] = _df_effect['Суммы_взыскано_по_иску'].fillna(0)    

_df_effect['SurchargeValue_cumsum_by_incident'] = _df_effect['SurchargeValue_cumsum_by_incident'].fillna(0)    
_df_effect['UTSSurchargeValue_cumsum_by_incident'] = _df_effect['UTSSurchargeValue_cumsum_by_incident'].fillna(0)

_df_effect['fin_effect_fact'] = (_df_effect['Сумма_выплат_по_претензиям'] + \
                                 # _df_effect['SurchargeValue_cumsum_by_incident'] + \
                                 # _df_effect['UTSSurchargeValue_cumsum_by_incident'] + \
                                 _df_effect['Сумма_взыскано_по_ФУ'] + \
                                 _df_effect['Суммы_взыскано_по_иску'] + \
                                 _df_effect['Взносы'])

display(_df_effect['TARGET'].value_counts())

_df_effect.loc[(_df_effect['TARGET'] == '0') & (_df_effect['Сумма_выплат_по_претензиям'] > 0), 'TARGET'] = '1'

display(_df_effect['TARGET'].value_counts())

>> Гатятулин Чингиз Русланович [01.07.2026 18:04]
Расчет фин. эффекта (новый)¶
>> Гатятулин Чингиз Русланович [01.07.2026 18:04]
_df_effect = X_test_freq.copy()
_df_effect.shape
>> Гатятулин Чингиз Русланович [01.07.2026 18:04]
def payments_fee(row):
    payments = 0
    if row['Сумма_взыскано_по_ФУ'] > 0:
        payments += 100000
#    if row['Суммы_взыскано_по_иску'] > 0:
#        payments += 15000

    return payments

_df_effect['Взносы'] = _df_effect.apply(payments_fee, axis=1)
_df_effect['Взносы'].value_counts(dropna=False)
>> Гатятулин Чингиз Русланович [01.07.2026 18:04]
_df_effect['Сумма_выплат_по_претензиям'] = _df_effect['Сумма_выплат_по_претензиям'].fillna(0)
_df_effect['Сумма_взыскано_по_ФУ'] = _df_effect['Сумма_взыскано_по_ФУ'].fillna(0)
_df_effect['Суммы_взыскано_по_иску'] = _df_effect['Суммы_взыскано_по_иску'].fillna(0)    

_df_effect['SurchargeValue_cumsum_by_incident'] = _df_effect['SurchargeValue_cumsum_by_incident'].fillna(0)    
_df_effect['UTSSurchargeValue_cumsum_by_incident'] = _df_effect['UTSSurchargeValue_cumsum_by_incident'].fillna(0)
>> Гатятулин Чингиз Русланович [01.07.2026 18:04]
_df_effect['fin_effect_fact'] = (_df_effect['Сумма_выплат_по_претензиям'] + \
                                 # _df_effect['SurchargeValue_cumsum_by_incident'] + \
                                 # _df_effect['UTSSurchargeValue_cumsum_by_incident'] + \
                                 _df_effect['Сумма_взыскано_по_ФУ'] + \
                                 _df_effect['Суммы_взыскано_по_иску'] + \
                                 _df_effect['Взносы'])
>> Гатятулин Чингиз Русланович [01.07.2026 18:04]
display(_df_effect['TARGET'].value_counts())

_df_effect.loc[(_df_effect['TARGET'] == '0') & (_df_effect['Сумма_выплат_по_претензиям'] > 0), 'TARGET'] = '1'

display(_df_effect['TARGET'].value_counts())
>> Гатятулин Чингиз Русланович [01.07.2026 18:04]
import numpy as np
import pandas as pd
from tqdm import tqdm

# ============================================================================
# 0. Подготовка: _df_effect теперь полного размера
# ============================================================================
# _df_effect = X_test_freq.copy()  # ← вы уже сделали это

# Но: для строк без иска сумма взыскания = 0, и выплаты по претензиям/ФУ тоже могут быть 0 — это нормально.
# Однако, если в ваших данных "Суммы_взыскано_по_иску" и т.п. уже содержат 0 там, где нет иска — всё корректно.

# ============================================================================
# 1. Предсказания
# ============================================================================

y_proba_freq = model_freq.predict_proba(_df_effect[summary_freq['selected_features_names']])[:, 1]
y_pred_sev = model_sev.predict(_df_effect[summary_sev['selected_features_names']])

# ============================================================================
# 2. Целевые переменные (должны быть длины N)
# ============================================================================

# y_test_freq: бинарный вектор (0/1), длина N
# y_test_sev: суммы только для исков (длина M) → расширяем до N
y_true_sev_full = np.array(_df_effect['TARGET_3_SEV'])

# ============================================================================
# 3. Подбор порога
# ============================================================================

thresholds = np.arange(0.0, 1.1, 0.1)
results = {}

base_sum = (
    _df_effect['fin_effect_fact'].values
)

for th in tqdm(thresholds, desc="Подбор порога классификации"):
    pred_freq = (y_proba_freq >= th).astype(int)
    
    mask_00 = (pred_freq == 0) & (y_test_freq == 0)
    mask_01 = (pred_freq == 0) & (y_test_freq == 1)
    mask_10 = (pred_freq == 1) & (y_test_freq == 0)
    mask_11 = (pred_freq == 1) & (y_test_freq == 1)
    
    fin_effect_model = np.zeros(len(_df_effect))
    
    fin_effect_model[mask_00] = -base_sum[mask_00]
    fin_effect_model[mask_01] = -base_sum[mask_01]

    fin_effect_model[mask_10] = -y_pred_sev[mask_10] - base_sum[mask_10]
    
    mask_11_over = mask_11 & (y_pred_sev >= y_true_sev_full)
    mask_11_under = mask_11 & (y_pred_sev < y_true_sev_full)
    fin_effect_model[mask_11_over] = -y_pred_sev[mask_11_over]
    fin_effect_model[mask_11_under] = -base_sum[mask_11_under]
    
    total_effect_model = fin_effect_model.sum()
    total_effect_fact = base_sum.sum()
    net_effect = total_effect_model - (-total_effect_fact)
    
    tp = np.sum((pred_freq == 1) & (y_test_freq == 1))
    n_pred = pred_freq.sum()
    n_actual = y_test_freq.sum()
    
    results[round(th, 2)] = {
        'net_effect': net_effect,
        'total_model': total_effect_model,
        'total_fact': total_effect_fact,
        'n_positive_preds': n_pred,
        'n_actual_positive': n_actual,
        'precision': tp / max(n_pred, 1),
        'recall': tp / max(n_actual, 1)
    }

# ============================================================================
# 4. Лучший порог
# ============================================================================

best_threshold = max(results, key=lambda k: results[k]['net_effect'])
best = results[best_threshold]

print("\n" + "="*70)
print("✅ ОПТИМАЛЬНЫЙ ПОРОГ КЛАССИФИКАЦИИ")
print("="*70)
print(f"Порог вероятности       : {best_threshold:.2f}")
print(f"Чистый финансовый эффект: {best['net_effect']:,.2f} ₽")

# ============================================================================
# 5. Сохранение предсказаний в _df_effect
# ============================================================================

best_pred_freq = (y_proba_freq >= best_threshold).astype(int)

_df_effect['pred_freq'] = best_pred_freq
_df_effect['pred_sev'] = y_pred_sev

# Финальный расчёт fin_effect_model
mask_00 = (best_pred_freq == 0) & (y_test_freq == 0)
mask_01 =(best_pred_freq == 0) & (y_test_freq == 1)
mask_10 = (best_pred_freq == 1) & (y_test_freq == 0)
mask_11 = (best_pred_freq == 1) & (y_test_freq == 1)

_df_effect['fin_effect_model'] = 0.0
_df_effect.loc[mask_00, 'fin_effect_model'] = -base_sum[mask_00]
_df_effect.loc[mask_01, 'fin_effect_model'] = -base_sum[mask_01]
_df_effect.loc[mask_10, 'fin_effect_model'] = -y_pred_sev[mask_10] - base_sum[mask_10]
_df_effect.loc[mask_11 & (y_pred_sev >= y_true_sev_full), 'fin_effect_model'] = -y_pred_sev[mask_11 & (y_pred_sev >= y_true_sev_full)]
_df_effect.loc[mask_11 & (y_pred_sev < y_true_sev_full), 'fin_effect_model'] = -base_sum[mask_11 & (y_pred_sev < y_true_sev_full)]

# _df_effect['fin_effect_fact'] = -base_sum

print(f"\nГотово! Колонки добавлены в _df_effect:")
print(f"  • pred_freq       — предсказание наличия иска (0/1)")
print(f"  • pred_sev        — прогнозная сумма взыскания")
print(f"  • fin_effect_model — финансовый эффект по модели")
print(f"\nИтоговый чистый эффект: {(_df_effect['fin_effect_model'].sum() - (-_df_effect['fin_effect_fact'].sum())):,.2f} ₽")

_df_effect['fin_effect_model'].sum()

_df_effect['fin_effect_fact'].sum()

_df_effect['pred_sev'].sum()

import pandas as pd
from IPython.display import display
import numpy as np

def create_summary_table(_df_effect):
    """Создает сводную таблицу по всем комбинациям TARGET_2 и pred_freq"""
    
    masks = {
        '1_1': (_df_effect['TARGET_2'] == 1) & (_df_effect['pred_freq'] == 1),
        '1_0': (_df_effect['TARGET_2'] == 1) & (_df_effect['pred_freq'] == 0),
        '0_1': (_df_effect['TARGET_2'] == 0) & (_df_effect['pred_freq'] == 1),
        '0_0': (_df_effect['TARGET_2'] == 0) & (_df_effect['pred_freq'] == 0),
    }
    
    rows = []
    
    for mask_name, mask in masks.items():
        df_group = _df_effect[mask]
        target_2, pred_freq = map(int, mask_name.split('_'))
        
        # Считаем все метрики (делаем отрицательными)
        count = df_group.shape[0]
        payout_main = -df_group["Выплата_по_основному_убытку"].sum()  # Отрицательное
        sum_od_uts = -df_group["TARGET_3_SEV"].sum()  # Отрицательное
        regression = -df_group["pred_sev"].sum()  # Отрицательное
        sum_claims = -df_group["Сумма_выплат_по_претензиям"].sum()  # Отрицательное
        sum_fu = -df_group["Сумма_взыскано_по_ФУ"].sum()  # Отрицательное
        sum_lawsuit = -df_group["Суммы_взыскано_по_иску"].sum()  # Отрицательное
        contributions = -df_group["Взносы"].sum()  # Отрицательное
        
        # Финансовые эффекты (отрицательные)
        fin_effect_model = df_group['fin_effect_model'].sum()  # Отрицательное
        fin_effect_fact = sum_claims + sum_fu + sum_lawsuit + contributions  # Уже отрицательное
        
        # Определяем ИТОГО в зависимости от группы
        if target_2 == 1 and pred_freq == 1:
            # 1/1 (TP): ИТОГО = ФИН. ЭФФЕКТ МОДЕЛЬ
            total = fin_effect_model
        elif target_2 == 1 and pred_freq == 0:
            # 1/0 (FN): ИТОГО = ФИН. ЭФФЕКТ ФАКТ
            total = fin_effect_fact
        elif target_2 == 0 and pred_freq == 1:
            # 0/1 (FP): ИТОГО = ФИН. ЭФФЕКТ МОДЕЛЬ - ФИН. ЭФФЕКТ ФАКТ
            total = fin_effect_model - fin_effect_fact
        else:
            # 0/0 (TN): ИТОГО = 0 (или фин. эффект факт, но он должен быть 0)
            total = 0
        
        rows.append({
            'Количество инцидентов с иными взысканиями': count,
            'Факт': target_2,
            'Классификация': pred_freq,
            'Выплата по основному убытку': payout_main,
            'Сумма ОД+УТС+Износ': sum_od_uts,
            'Регрессия': regression,
            'Сумма выплат по претензиям': sum_claims,
            'Сумма взыскано по ФУ': sum_fu,
            'Суммы взыскано по иску': sum_lawsuit,
            'Взносы': contributions,
            'ФИН. ЭФФЕКТ МОДЕЛЬ': fin_effect_model,
            'ФИН. ЭФФЕКТ ФАКТ': fin_effect_fact,
            'ИТОГО': total
        })
    
    return pd.DataFrame(rows)

def color_excel_table(writer, sheet_name, df):
    """Раскрашивает Excel таблицу в цвета как на скриншоте"""
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    
    workbook = writer.book
    worksheet = workbook[sheet_name]
    
    # Цвета из скриншота
    colors = {
        'gray': '4F4F4F',        # Темно-серый (Количество, Факт, ИТОГО)
        'purple': '8B479C',      # Фиолетовый (Классификация, Регрессия)
        'yellow': 'FFC000',      # Желтый (Выплата по основному убытку)
        'blue': '00B0F0',        # Голубой (Сумма ОД+УТС+Износ)
        'pink': 'FF9999',        # Розовый (Сумма выплат, взыскано)
        'green': '00B050',       # Зеленый (ФИН. ЭФФЕКТ МОДЕЛЬ)
        'red': 'FF0000',         # Красный (ФИН. ЭФФЕКТ ФАКТ)
        'white': 'FFFFFF'        # Белый текст
    }
    
    # Соответствие столбцов и цветов
    column_colors = {
        'Количество инцидентов с иными взысканиями': colors['gray'],
        'Факт': colors['gray'],
        'Классификация': colors['purple'],
'Вып ла та по ос но вн ом у у бы тк у' : c ol or s[ 'y el lo w' ] ,
        'Сумма ОД+УТС+Износ': colors['blue'],
        'Регрессия': colors['purple'],
        'Сумма выплат по претензиям': colors['pink'],
        'Сумма взыскано по ФУ': colors['pink'],
        'Суммы взыскано по иску': colors['pink'],
        'Взносы': colors['pink'],
        'ФИН. ЭФФЕКТ МОДЕЛЬ': colors['green'],
        'ФИН. ЭФФЕКТ ФАКТ': colors['red'],
        'ИТОГО': colors['gray']
    }
    
    # Создаем форматирование
    header_font = Font(bold=True, color=colors['white'])
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(horizontal='right', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Форматируем заголовки
    for col_num, cell in enumerate(worksheet[1], 1):
        col_name = df.columns[col_num - 1] if col_num - 1 < len(df.columns) else None
        if col_name and col_name in column_colors:
            fill_color = column_colors[col_name]
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
    
    # Форматируем данные
    for row_num in range(2, worksheet.max_row + 1):
        for col_num, cell in enumerate(worksheet[row_num], 1):
            col_name = df.columns[col_num - 1] if col_num - 1 < len(df.columns) else None
            if col_name and col_name in column_colors:
                fill_color = column_colors[col_name]
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                cell.alignment = cell_alignment
                cell.border = thin_border
                
                # Форматируем числа
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
    
    # Устанавливаем ширину столбцов
    column_widths = {
        'Количество инцидентов с иными взысканиями': 25,
        'Факт': 8,
        'Классификация': 15,
        'Выплата по основному убытку': 20,
        'Сумма ОД+УТС+Износ': 20,
        'Регрессия': 15,
        'Сумма выплат по претензиям': 22,
        'Сумма взыскано по ФУ': 18,
        'Суммы взыскано по иску': 20,
        'Взносы': 12,
        'ФИН. ЭФФЕКТ МОДЕЛЬ': 18,
        'ФИН. ЭФФЕКТ ФАКТ': 18,
        'ИТОГО': 15
    }
    
    for col_num, col_name in enumerate(df.columns, 1):
        if col_name in column_widths:
            worksheet.column_dimensions[chr(64 + col_num) if col_num <= 26 else chr(64 + col_num - 26) + chr(65)].width = column_widths[col_name]
        else:
            worksheet.column_dimensions[chr(64 + col_num)].width = 15

# Создаем таблицу
summary_table = create_summary_table(_df_effect)

# Отображаем в Jupyter
display(summary_table.style.format({
    'Количество инцидентов с иными взысканиями': '{:,.0f}'.format,
    'Выплата по основному убытку': '{:,.0f}'.format,
    'Сумма ОД+УТС+Износ': '{:,.0f}'.format,
    'Регрессия': '{:,.0f}'.format,
    'Сумма выплат по претензиям': '{:,.0f}'.format,
    'Сумма взыскано по ФУ': '{:,.0f}'.format,
    'Суммы взыскано по иску': '{:,.0f}'.format,
    'Взносы': '{:,.0f}'.format,
    'ФИН. ЭФФЕКТ МОДЕЛЬ': '{:,.0f}'.format,
    'ФИН. ЭФФЕКТ ФАКТ': '{:,.0f}'.format,
    'ИТОГО': '{:,.0f}'.format,
}))

# Экспортируем в Excel с форматированием
output_file = 'summary_table_3.xlsx'
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    summary_table.to_excel(writer, sheet_name='Summary', index=False)
    color_excel_table(writer, 'Summary', summary_table)

print(f"Таблица сохранена в файл: {output_file}")

# Выводим информацию о логике расчета
print("\n===ЛОГИКА РАСЧЕТА ФИНАНСОВЫХ ЭФФЕКТОВ ===")
print("1/1 (TP): ИТОГО = ФИН. ЭФФЕКТ МОДЕЛЬ")
print("1/0 (FN): ИТОГО = ФИН. ЭФФЕКТ ФАКТ")
print("0/1 (FP): ИТОГО = ФИН. ЭФФЕКТ МОДЕЛЬ - ФИН. ЭФФЕКТ ФАКТ")
print("0/0 (TN): ИТОГО = 0")
print("\nВсе суммы отрицательные (расходы)")

_df_effect.shape

_df_effect[_df_effect['TARGET'] == '1'].shape[0] / _df_effect.shape[0] * 100

(_df_effect[_df_effect['TARGET'] == '1'].shape[0] -
_df_effect[(_df_effect['TARGET'] == '1') & (_df_effect['pred_freq'] == 1)].shape[0]) / _df_effect.shape[0] * 100

# Средняя выплата увеличится на:

_df_effect['Выплата_по_основному_убытку'].median()

_df_effect['pred_sev'].median()

_df_effect.shape

import matplotlib.pyplot as plt
import numpy as np
from sklearn.metrics import confusion_matrix, ConfusionMatrixDisplay

# Пример истинных и предсказанных меток
y_true = _df_effect['TARGET_2'].astype(int)
y_pred = _df_effect['pred_freq']

# Вычисление матрицы ошибок
cm = confusion_matrix(y_true, y_pred)

# Визуализация
# Создаем объект без параметров форматирования
disp = ConfusionMatrixDisplay(confusion_matrix=cm)

# Передаем values_format прямо в метод plot()
disp.plot(cmap=plt.cm.Blues, values_format='d') 

plt.title("Confusion Matrix")
plt.show()


import numpy as np
import matplotlib.pyplot as plt
from sklearn.metrics import precision_score, recall_score
from catboost import CatBoostClassifier

def plot_precision_recall_vs_threshold(model, X, y_true, thresholds=None, ax=None):
    """
    Построение графика зависимости Precision и Recall от порога классификации.
    
    Parameters:
    -----------
    model : CatBoostClassifier
        Обученная модель CatBoost
    X : array-like
        Признаки для предсказания
    y_true : array-like
        Истинные метки классов
    thresholds : array-like, optional
        Массив порогов для проверки. По умолчанию от 0 до 1 с шагом 0.01
    ax : matplotlib.axes.Axes, optional
        Ось для отрисовки (для использования в subplot)
    
    Returns:
    --------
    fig, ax : matplotlib фигуры и оси
    """
    # Получаем вероятности положительного класса
    y_proba = model.predict_proba(X)[:, 1]
    
    # Создаем массив порогов
    if thresholds is None:
        thresholds = np.linspace(0, 1, 101)
    
    # Вычисляем Precision и Recall для каждого порога
    precisions = []
    recalls = []
    
    for threshold in thresholds:
        y_pred = (y_proba >= threshold).astype(int)
        
        # Обработка случая, когда нет положительных предсказаний
        if y_pred.sum() == 0:
            precision = 1.0
        else:
            precision = precision_score(y_true, y_pred, zero_division=1)
        
        recall = recall_score(y_true, y_pred, zero_division=0)
        
        precisions.append(precision)
        recalls.append(recall)
    
    # Создаем фигуру и ось, если не переданы
    if ax is None:
        fig, ax = plt.subplots(figsize=(10, 6))
    else:
        fig = ax.figure
    
    # Построение графика
    ax.plot(thresholds, precisions, 'b-', label='Precision', linewidth=2)
    ax.plot(thresholds, recalls, 'r-', label='Recall', linewidth=2)
    ax.axhline(y=0.5, color='gray', linestyle='--', alpha=0.5)
    ax.axvline(x=0.5, color='gray', linestyle='--', alpha=0.5)
    
    # Настройка оформления
    ax.set_xlabel('Threshold', fontsize=12)
    ax.set_ylabel('Score', fontsize=12)
    ax.set_title('Precision и Recall в зависимости от порога классификации', fontsize=14)
    ax.legend(loc='best', fontsize=11)
    ax.grid(True, alpha=0.3)
    ax.set_xlim([0, 1])
    ax.set_ylim([0, 1.05])
    
    # Добавляем информацию о лучшем пороге (F1-score)
    f1_scores = []
    for p, r in zip(precisions, recalls):
        if p + r > 0:
            f1_scores.append(2 * p * r / (p + r))
        else:
            f1_scores.append(0)
    
    best_threshold_idx = np.argmax(f1_scores)
    best_threshold = thresholds[best_threshold_idx]
    best_f1 = f1_scores[best_threshold_idx]
    
    ax.axvline(x=best_threshold, color='green', linestyle='--', 
               label=f'Лучший threshold (F1={best_f1:.3f})', linewidth=2)
    
    plt.tight_layout()
    
    return fig, ax, thresholds, precisions, recalls


# Построение графика
fig, ax, thresholds, precisions, recalls = plot_precision_recall_vs_threshold(
    model_freq, X_test_freq[summary_freq['selected_features_names']], y_test_freq
)

plt.show()

# Вывод лучших порогов
print(f"Лучший порог (по F1): {thresholds[np.argmax([2*p*r/(p+r) if p+r > 0 else 0 for p, r in zip(precisions, recalls)])]:.2f}")

import matplotlib.pyplot as plt
import seaborn as sns
import pandas as pd

# ==============================================================================
# 1. Подготовка масок квадрантов Confusion Matrix
# ==============================================================================
y_true = y_test_freq  # Факт наличия иска (0/1)
y_pred = _df_effect['pred_freq']  # Прогноз модели (0/1)

tn_mask = (y_true == 0) & (y_pred == 0)  # True Negative
fp_mask = (y_true == 0) & (y_pred == 1)  # False Positive
fn_mask = (y_true == 1) & (y_pred == 0)  # False Negative
tp_mask = (y_true == 1) & (y_pred == 1)  # True Positive

# ==============================================================================
# 2. Агрегация сумм расходов (берем модуль, т.к. в эффекте они отрицательные)
# ==============================================================================
def get_costs(mask, col_name):
    return abs(_df_effect.loc[mask, col_name].sum())

fact_costs = [
    [get_costs(tn_mask, 'fin_effect_fact'), get_costs(fp_mask, 'fin_effect_fact')],
    [get_costs(fn_mask, 'fin_effect_fact'), get_costs(tp_mask, 'fin_effect_fact')]
]

model_costs = [
    [get_costs(tn_mask, 'fin_effect_model'), get_costs(fp_mask, 'fin_effect_model')],
    [get_costs(fn_mask, 'fin_effect_model'), get_costs(tp_mask, 'fin_effect_model')]
]

# ==============================================================================
# 3. Визуализация
# ==============================================================================
index = ['Нет иска (0)', 'Иск (1)']
columns = ['Прогноз 0', 'Прогноз 1']

df_fact = pd.DataFrame(fact_costs, index=index, columns=columns)
df_model = pd.DataFrame(model_costs, index=index, columns=columns)

fig, axes = plt.subplots(1, 2, figsize=(14, 6))

# heatmap факта
sns.heatmap(df_fact, annot=True, fmt=',.0f', cmap='Blues', ax=axes[0], 
            cbar_kws={'label': 'Сумма (₽)'}, annot_kws={'size': 11})
axes[0].set_title('Фактические судебные расходы', fontsize=14, weight='bold')

# heatmap модели
sns.heatmap(df_model, annot=True, fmt=',.0f', cmap='Greens', ax=axes[1], 
            cbar_kws={'label': 'Сумма (₽)'}, annot_kws={'size': 11})
axes[1].set_title('Расходы по модели (Частота + Тяжесть)', fontsize=14, weight='bold')

plt.tight_layout()
plt.show()

%matplotlib inline

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt

# Попытка импорта plotly
try:
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots
    PLOTLY_AVAILABLE = True
except ImportError:
    PLOTLY_AVAILABLE = False
    print("Plotly не установлен. Используется matplotlib.")

# ============================================================================
# 1. Подготовка данных ТОЛЬКО для строк с реальным иском
# ============================================================================

mask_actual_claim = (y_test_freq == 1)

plot_data = pd.DataFrame({
    'fact_sev': y_true_sev_full[mask_actual_claim],
    'pred_sev': y_pred_sev[mask_actual_claim],
    'base_sum': base_sum[mask_actual_claim],
    'fact_surcharge': _df_effect[mask_actual_claim]['Выплата_по_основному_убытку']
})

# Удаляем некорректные значения
plot_data = plot_data[(plot_data['fact_sev'] > 0) & (plot_data['pred_sev'] >= 0)]

# Определяем границы бинов — например, до 2 млн руб с шагом 100 тыс.
max_val = plot_data['fact_sev'].quantile(0.95)  # обрезаем выбросы (95-й перцентиль)
bins = np.arange(0, max_val + 100_000, 100_000)  # шаг 100 тыс. руб

# Используем pd.cut вместо qcut
plot_data['fact_sev_bin'] = pd.cut(
    plot_data['fact_sev'],
    bins=25,
    include_lowest=True,
    right=False  # [a, b)
).apply(lambda x: x.left if pd.notna(x) else np.nan)

# Удаляем строки, не попавшие в бины (если есть)
plot_data = plot_data.dropna(subset=['fact_sev_bin'])
plot_data['fact_sev_bin'] = plot_data['fact_sev_bin'].astype(float)

# Группировка
GB = plot_data.groupby('fact_sev_bin').agg(
    fact_sev_median=('fact_sev', 'median'),
    pred_sev=('pred_sev', 'median'),
    base_sum=('base_sum', 'median'),
    fact_surcharge=('fact_surcharge', 'median'),
    n_claims=('fact_sev', 'count')
).reset_index()

# ============================================================================
# 2. Построение графика
# ============================================================================

if PLOTLY_AVAILABLE:
    fig = make_subplots(specs=[[{"secondary_y": True}]])

    # Линии на левой оси
    fig.add_trace(go.Scatter(
        x=GB['fact_sev_bin'],
        y=GB['fact_sev_median'] + GB['fact_surcharge'],
        mode='lines+markers',
        name='Основной долг + УТС + Износ',
        line=dict(color='#FF6B6B', width=3),
        marker=dict(symbol='circle', size=8, color='#FF6B6B', line=dict(width=1, color='white')),
        hovertemplate="<b>Основной долг + УТС + Износ</b><br>Сумма: %{y:,.0f} ₽<extra></extra>"
    ), secondary_y=False)

    fig.add_trace(go.Scatter(
        x=GB['fact_sev_bin'],
        y=GB['pred_sev'] + GB['fact_surcharge'],
        mode='lines+markers',
        name='Прогноз модели',
        line=dict(color='#4ECDC4', width=3),
        marker=dict(symbol='diamond', size=8, color='#4ECDC4', line=dict(width=1, color='white')),
        hovertemplate="<b>Прогноз модели</b><br>Сумма: %{y:,.0f} ₽<extra></extra>"
    ), secondary_y=False)

    fig.add_trace(go.Scatter(
        x=GB['fact_sev_bin'],
        y=GB['base_sum'] + GB['fact_surcharge'],
        mode='lines+markers',
        name='ПСР',
        line=dict(color='#9B5DE5', width=3),
        marker=dict(symbol='square', size=8, color='#9B5DE5', line=dict(width=1, color='white')),
        hovertemplate="<b>Выплаты</b><br>Сумма: %{y:,.0f} ₽<extra></extra>"
    ), secondary_y=False)

    fig.add_trace(go.Scatter(
        x=GB['fact_sev_bin'],
        y=GB['fact_surcharge'],
        mode='lines+markers',
        name='Фактические выплаты по убытку',
        line=dict(color='silver', width=3),
        marker=dict(symbol='square', size=8, color='silver', line=dict(width=1, color='white')),
        hovertemplate="<b>Фактические выплаты</b><br>Сумма: %{y:,.0f} ₽<extra></extra>"
    ), secondary_y=False)

    # Столбцы на правой оси
fig.add_trace(go.Bar(
        x=GB['fact_sev_bin'],
        y=GB['n_claims'],
        name='Количество исков',
        marker_color='rgba(100, 100, 100, 0.3)',
        opacity=0.6,
        hovertemplate="<b>Экспозиция</b><br>Иски: %{y}<extra></extra>"
    ), secondary_y=True)

    # Настройка осей
    fig.update_xaxes(
        title="Основной долг + УТС + Износ (бин, руб)",
        title_font=dict(size=14),
        tickfont=dict(size=12),
        gridcolor='lightgray',
        tickformat=',.0f'
    )

    fig.update_yaxes(
        title="Сумма (руб)",
        title_font=dict(size=14, color='#333'),
        tickfont=dict(size=12),
        gridcolor ='lightgray' ,
        tickformat=',.0f',
        secondary_y=False
    )

    fig.update_yaxes(
        title="Количество исков",
        title_font=dict(size=14, color='gray'),
        tickfont=dict(size=12, color='gray'),
        showgrid=False,
        secondary_y=True
    )

    fig.update_layout(
        title=dict(
            text="Прогноз vs Факт: суммы и экспозиция (равномерные бины)\n",
            x=0.5,
            font=dict(size=18)
        ),
        plot_bgcolor='white',
        paper_bgcolor='white',
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="right",
            x=1,
            bgcolor="rgba(255,255,255,0.8)"
        ),
        width=1150,
        height=800,
        margin=dict(l=80, r=80, t=80, b=80),
        hovermode='x unified'
    )

    fig.show()

else:
    # Matplotlib fallback
    fig, ax1 = plt.subplots(figsize=(15, 15))

    ax1.plot(GB['fact_sev_bin'], GB['fact_sev_median'], 
             color='#FF6B6B', marker='o', linewidth=2, markersize=6, label='Основной долг + УТС + Износ')
    ax1.plot(GB['fact_sev_bin'], GB['pred_sev'], 
             color='#4ECDC4', marker='D', linewidth=2, markersize=6, label='Прогноз модели')
    ax1.plot(GB['fact_sev_bin'], GB['base_sum'], 
             color='#9B5DE5', marker='s', linewidth=2, markersize=6, label='ПСР')

    ax1.set_xlabel("Сумма (бин, руб)", fontsize=12)
    ax1.set_ylabel("Сумма (руб)", fontsize=12)
    ax1.ticklabel_format(style='plain', axis='y')
    ax1.grid(True, alpha=0.3)

    ax2 = ax1.twinx()
    bin_width = np.diff(GB['fact_sev_bin']).mean() if len(GB) > 1 else 100_000
    ax2.bar(GB['fact_sev_bin'], GB['n_claims'], 
            color='gray', alpha=0.3, width=bin_width * 0.8, label='Количество исков')
    ax2.set_ylabel("Количество исков", color='gray')
    ax2.tick_params(axis='y', labelcolor='gray')

    lines_1, labels_1 = ax1.get_legend_handles_labels()
    lines_2, labels_2 = ax2.get_legend_handles_labels()
    ax1.legend(lines_1 + lines_2, labels_1 + labels_2, loc='upper left', bbox_to_anchor=(0, -0.15), ncol=2)

#    plt.title("Прогноз vs Факт: суммы и экспозиция (равномерные бины)", fontsize=14, pad=20)
    plt.tight_layout(rect=[0, 0.05, 1, 1])
    plt.show()
    
    
import plotly.express as px
import plotly.io as pio

# Сохраняем как интерактивный HTML
pio.write_html(fig, file="my_plot.html", auto_open=True)


_df_effect['fin_effect_fact'] = -_df_effect['fin_effect_fact']


rename_dict = {
    'INCIDENT_NUMBER': 'НОМЕР ИНЦИДЕНТА',
    'FILIAL': 'ФИЛИАЛ',
    'Выплата_по_основному_убытку': 'ВЫПЛАТА ПО ОСНОВНОМУ УБЫТКУ',
    'Сумма_выплат_по_претензиям': 'СУММА ВЫПЛАТ ПО ПРЕТЕНЗИЯМ',
    'Сумма_взыскано_по_ФУ': 'СУММА ВЗЫСКАННАЯ У ФУ',
    'Суммы_взыскано_по_иску': 'СУММА ВЗЫСКАННАЯ В СУДЕ',
    'Взносы': 'ВЗНОСЫ',
    'fin_effect_fact': 'ФАКТ ФИН. ЭФФЕКТ ',
    'TARGET_3_FREQ': 'ФАКТ БЫЛО ВЗЫСКАНИЕ ОСНОВНОГО ДОЛГА/УТС/ИЗНОСА',
    'TARGET_3_SEV': 'ФАКТ СУММА ВЗЫСКАНИЯ ОСНОВНОГО ДОЛГА/УТС/ИЗНОСА',
    'TARGET_2': 'БЫЛ ПСР (НОВЫЙ ТАРГЕТ)',
    'pred_freq': 'МОДЕЛЬ БУДЕТ ЛИ ВЗЫСКАНИЕ ОСНОВНОГО ДОЛГА/УТС/ИЗНОСА',
    'pred_sev': 'МОДЕЛЬ СУММА ВЗЫСКАНИЯ ОСНОВНОГО ДОЛГА/УТС/ИЗНОСА',
    'fin_effect_model': 'МОДЕЛЬ ФИН. ЭФФЕКТ',
}

_df_effect[['INCIDENT_NUMBER', 'FILIAL', 'Выплата_по_основному_убытку',
            'Сумма_выплат_по_претензиям', 'Сумма_взыскано_по_ФУ', 'Суммы_взыскано_по_иску', 'Взносы',
            'fin_effect_fact', 'TARGET_3_FREQ', 'TARGET_3_SEV', 'TARGET_2',
            'pred_freq', 'pred_sev', 'fin_effect_model']].rename(rename_dict, axis=1).to_excel('Аналитика.xlsx')

_df_effect[['INCIDENT_NUMBER', 'FILIAL', 'Выплата_по_основному_убытку',
            'Сумма_выплат_по_претензиям', 'Сумма_взыскано_по_ФУ', 'Суммы_взыскано_по_иску', 'Взносы',
            'fin_effect_fact', 'TARGET_3_FREQ', 'TARGET_3_SEV', 'TARGET_2',
            'pred_freq', 'pred_sev', 'fin_effect_model']].rename(rename_dict, axis=1)

import pandas as pd
import matplotlib.pyplot as plt

# Приводим дату к datetime
df['ORDER_DATETIME'] = pd.to_datetime(df['PAYMENT_ORDER_DATE_TIME'])

# Группируем по месяцам и считаем долю единиц
monthly_share = df.groupby(df['ORDER_DATETIME'].dt.to_period('M')).agg(
    target_share=('TARGET_2', 'mean'),      # доля положительных
    total_count=('TARGET_2', 'count')       # общий объём
).reset_index()

# Преобразуем период в дату для отображения
monthly_share['MONTH_START'] = monthly_share['ORDER_DATETIME'].dt.start_time

# Строим bar-чарт
plt.figure(figsize=(12, 5))
plt.bar(monthly_share['MONTH_START'], monthly_share['target_share'] * 100, width=25, alpha=0.8 )

plt.xlabel('Месяц')
plt.ylabel('Доля положительных случаев, %')
plt.title('Доля TARGET_2 = 1 в общем объёме (помесячно)')
plt.grid(axis='y', alpha=0.3, linestyle='--')
plt.xticks(rotation=45)
plt.tight_layout()
plt.show()


import pandas as pd
import matplotlib.pyplot as plt

# Приводим дату к datetime
df['ORDER_DATETIME'] = pd.to_datetime(df['PAYMENT_ORDER_DATE_TIME'])

# Добавляем колонку с месяцем (для группировки)
df['MONTH'] = df['ORDER_DATETIME'].dt.to_period('M')

# Считаем только положительные случаи (TARGET_2 == 1)
positive_by_month = df[df['TARGET_2'] == 1].groupby('MONTH').size().reset_index(name='positive_count')

# Преобразуем период обратно в дату для красивого отображения
positive_by_month['MONTH_START'] = positive_by_month['MONTH'].dt.start_time

# Строим график
plt.figure(figsize=(12, 5))
plt.bar(positive_by_month['MONTH_START'], positive_by_month['positive_count'], width=25, alpha=0.8)
plt.xlabel('Месяц')
plt.ylabel('Количество положительных случаев (TARGET_2 = 1)')
plt.title('Положительные случаи по месяцам')
plt.grid(axis='y', alpha=0.3)
plt.xticks(rotation=45)
plt.tight_layout()
plt.show()