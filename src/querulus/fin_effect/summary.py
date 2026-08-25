"""Сводная таблица финансового эффекта по квадрантам TARGET_FREQ × pred_freq.

Только агрегация колонок кадра после ``run_fin_effect_pipeline`` /
``apply_model_predictions``. Расчёт эффектов — в ``calculator`` (один контракт).
"""
from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd

from querulus.fin_effect.config import FinEffectConfig

_EFFECT_COLS = ("pred_freq", "fin_effect_model", "fin_effect_fact")


def _neg_column_sum(group: pd.DataFrame, column: str) -> float:
    """Сумма колонки с инверсией знака (расходы отрицательные)."""
    if column not in group.columns:
        return 0.0
    return float(-pd.to_numeric(group[column], errors="coerce").fillna(0).sum())


def _sum_col(group: pd.DataFrame, column: str) -> float:
    if column not in group.columns:
        return 0.0
    return float(pd.to_numeric(group[column], errors="coerce").fillna(0).sum())


def create_summary_table(
    effect_df: pd.DataFrame,
    config: FinEffectConfig | None = None,
    *,
    itogo_mode: str | None = None,
    verify: bool = True,
) -> pd.DataFrame:
    """Сводная таблица по комбинациям frequency_target и pred_freq.

    Берёт ``fin_effect_model`` / ``fin_effect_fact`` / ``fin_effect_economy``
    с кадра как есть (после calculator). Никакого пересчёта факта.

    ``Экономия`` = sum(``fin_effect_economy``) = model − fact по квадранту
    (как ``net_effect``: FP → отрицательная, удачное покрытие → положительная).
    ``Регрессия`` (pred_sev) — только при классификации = 1.
    ``itogo_mode`` — совместимость API, игнорируется.
    """
    config = config or FinEffectConfig()
    _ = itogo_mode
    missing = [col for col in _EFFECT_COLS if col not in effect_df.columns]
    if missing:
        raise ValueError(
            "create_summary_table: нет колонок "
            f"{missing}. Сначала run_fin_effect_pipeline / apply_model_predictions."
        )

    work = effect_df
    if "fin_effect_economy" not in work.columns:
        work = work.copy()
        work["fin_effect_economy"] = (
            pd.to_numeric(work["fin_effect_model"], errors="coerce").fillna(0)
            - pd.to_numeric(work["fin_effect_fact"], errors="coerce").fillna(0)
        )

    freq_col = (
        "fin_effect_y_true"
        if "fin_effect_y_true" in work.columns
        else config.frequency_target_column
    )
    if freq_col not in work.columns:
        raise ValueError(
            "create_summary_table: нет fin_effect_y_true / "
            f"{config.frequency_target_column}"
        )

    y_fact = pd.to_numeric(work[freq_col], errors="coerce").fillna(0).astype(int)
    y_pred = pd.to_numeric(work["pred_freq"], errors="coerce").fillna(0).astype(int)
    masks = {
        "1_1": (y_fact == 1) & (y_pred == 1),
        "1_0": (y_fact == 1) & (y_pred == 0),
        "0_1": (y_fact == 0) & (y_pred == 1),
        "0_0": (y_fact == 0) & (y_pred == 0),
    }

    rows: list[dict[str, float | int]] = []
    for mask_name, mask in masks.items():
        group = work.loc[mask]
        target_fact, pred_freq = map(int, mask_name.split("_"))

        regression = (
            _neg_column_sum(group, "pred_sev") if pred_freq == 1 else float("nan")
        )
        fin_effect_model = _sum_col(group, "fin_effect_model")
        fin_effect_fact = _sum_col(group, "fin_effect_fact")
        economy = _sum_col(group, "fin_effect_economy")

        rows.append(
            {
                "Количество инцидентов с иными взысканиями": int(group.shape[0]),
                "Факт": target_fact,
                "Классификация": pred_freq,
                "Выплата по основному убытку": _neg_column_sum(
                    group, config.base_payment_column
                ),
                "Сумма ОД+УТС+Износ": _neg_column_sum(
                    group, config.severity_target_column
                ),
                "Регрессия": regression,
                "Иски (TARGET_FREQ_CLAIMS)": _neg_column_sum(
                    group, config.freq_claims_amount_column
                ),
                "Претензии (TARGET_FREQ_PRET)": _neg_column_sum(
                    group, config.freq_pret_amount_column
                ),
                "Взносы": _neg_column_sum(group, config.premiums_column),
                "ФИН. ЭФФЕКТ МОДЕЛЬ": fin_effect_model,
                "ФИН. ЭФФЕКТ ФАКТ": fin_effect_fact,
                "Экономия": economy,
            }
        )

    summary = pd.DataFrame(rows)
    if verify:
        _verify_summary_matches_frame(summary, work)
    return summary


def _verify_summary_matches_frame(summary: pd.DataFrame, frame: pd.DataFrame) -> None:
    """Суммы квадрантов должны сходиться с итогами кадра."""
    checks = (
        ("ФИН. ЭФФЕКТ МОДЕЛЬ", "fin_effect_model"),
        ("ФИН. ЭФФЕКТ ФАКТ", "fin_effect_fact"),
        ("Экономия", "fin_effect_economy"),
    )
    for summary_col, frame_col in checks:
        left = float(summary[summary_col].sum())
        right = _sum_col(frame, frame_col)
        if abs(left - right) > 1.0:
            raise AssertionError(
                f"Сводка {summary_col}={left:,.2f} ≠ sum(frame.{frame_col})={right:,.2f}. "
                "Нарушен контракт calculator → summary."
            )


def compare_formula_summaries(
    effect_df: pd.DataFrame,
    config: FinEffectConfig | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Две сводки на тех же pred: старые квадранты модели и новые (coverage)."""
    from querulus.fin_effect.calculator import recompute_fin_effect_model

    config = config or FinEffectConfig()
    old_frame = effect_df.copy()
    old_frame["fin_effect_model"] = recompute_fin_effect_model(
        effect_df, config, formula="legacy"
    )
    old_frame["fin_effect_economy"] = (
        old_frame["fin_effect_model"] - old_frame["fin_effect_fact"]
    )
    new_frame = effect_df.copy()
    new_frame["fin_effect_model"] = recompute_fin_effect_model(
        effect_df, config, formula="coverage"
    )
    new_frame["fin_effect_economy"] = (
        new_frame["fin_effect_model"] - new_frame["fin_effect_fact"]
    )
    return (
        create_summary_table(old_frame, config),
        create_summary_table(new_frame, config),
    )


def color_excel_table(writer, sheet_name: str, summary_df: pd.DataFrame) -> None:
    """Раскрасить лист Excel как в Litigant."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    worksheet = writer.book[sheet_name]
    colors = {
        "gray": "4F4F4F",
        "purple": "8B479C",
        "yellow": "FFC000",
        "blue": "00B0F0",
        "pink": "FF9999",
        "green": "00B050",
        "red": "FF0000",
        "white": "FFFFFF",
    }
    column_colors = {
        "Количество инцидентов с иными взысканиями": colors["gray"],
        "Факт": colors["gray"],
        "Классификация": colors["purple"],
        "Выплата по основному убытку": colors["yellow"],
        "Сумма ОД+УТС+Износ": colors["blue"],
        "Регрессия": colors["purple"],
        "Иски (TARGET_FREQ_CLAIMS)": colors["pink"],
        "Претензии (TARGET_FREQ_PRET)": colors["pink"],
        "Взносы": colors["pink"],
        "ФИН. ЭФФЕКТ МОДЕЛЬ": colors["green"],
        "ФИН. ЭФФЕКТ ФАКТ": colors["red"],
        "Экономия": colors["gray"],
        "ИТОГО": colors["gray"],
    }
    header_font = Font(bold=True, color=colors["white"])
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_num, cell in enumerate(worksheet[1], 1):
        col_name = summary_df.columns[col_num - 1] if col_num - 1 < len(summary_df.columns) else None
        if col_name and col_name in column_colors:
            fill_color = column_colors[col_name]
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

    for row_num in range(2, worksheet.max_row + 1):
        for col_num, cell in enumerate(worksheet[row_num], 1):
            col_name = summary_df.columns[col_num - 1] if col_num - 1 < len(summary_df.columns) else None
            if col_name and col_name in column_colors:
                fill_color = column_colors[col_name]
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.alignment = cell_alignment
                cell.border = thin_border
                if isinstance(cell.value, (int, float)) and not (
                    isinstance(cell.value, float) and np.isnan(cell.value)
                ):
                    cell.number_format = "#,##0"

    column_widths = {
        "Количество инцидентов с иными взысканиями": 25,
        "Факт": 8,
        "Классификация": 15,
        "Выплата по основному убытку": 20,
        "Сумма ОД+УТС+Износ": 20,
        "Регрессия": 15,
        "Иски (TARGET_FREQ_CLAIMS)": 22,
        "Претензии (TARGET_FREQ_PRET)": 22,
        "Взносы": 12,
        "ФИН. ЭФФЕКТ МОДЕЛЬ": 18,
        "ФИН. ЭФФЕКТ ФАКТ": 18,
        "Экономия": 15,
        "ИТОГО": 15,
    }
    for col_num, col_name in enumerate(summary_df.columns, 1):
        width = column_widths.get(col_name, 15)
        worksheet.column_dimensions[get_column_letter(col_num)].width = width


def export_summary_excel(
    summary_df: pd.DataFrame,
    path: str | Path,
    *,
    sheet_name: str = "Summary",
) -> Path:
    """Сохранить сводную таблицу в Excel с форматированием."""
    output_path = Path(path)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name=sheet_name, index=False)
        color_excel_table(writer, sheet_name, summary_df)
    return output_path
